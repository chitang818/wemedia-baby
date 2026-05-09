"""
购物车推广 Excel 导入工具
文件路径：src/infrastructure/common/excel_cart_importer.py
功能：解析标准模板 Excel，将商品配置数据转换为字典列表，供仓储层批量导入。
"""

from __future__ import annotations

from typing import List, Dict, Any
import logging

from openpyxl import load_workbook

from src.domain.publish.promotion_limits import CART_SHORT_TITLE_MAX_LEN

logger = logging.getLogger(__name__)

# 必须存在的列
REQUIRED_HEADERS = ["商品简称"]
# 所有预期列（首行须全部包含）
EXPECTED_HEADERS = [
    "商品简称",
    "商品短标题",
    "抖音（链接）",
    "快手（商品名称）",
    "视频号（ID或链接）",
    "小红书（链接）",
]

# Excel 列名 -> 字典键映射
_COL_MAP = {
    "商品简称": "short_name",
    "商品短标题": "short_title",
    "抖音（链接）": "douyin_link",
    "快手（商品名称）": "kuaishou_product_name",
    "视频号（ID或链接）": "channels_id_or_link",
    "小红书（链接）": "xiaohongshu_link",
}


def parse_excel(path: str) -> Dict[str, Any]:
    """解析购物车推广 Excel 文件。

    Returns:
        dict: {
            "items": List[Dict[str, Any]],
            "total": int,
            "success": int,
            "failed": int,
            "errors": List[str],
        }
    """
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = wb.active

    headers: List[str] = []
    items: List[Dict[str, Any]] = []
    errors: List[str] = []
    total = 0
    success = 0

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [str(c or "").strip() for c in row]
            missing = [h for h in REQUIRED_HEADERS if h not in headers]
            if missing:
                raise ValueError(f"Excel 模板表头缺失：{', '.join(missing)}")
            continue

        if not any(row):
            continue

        total += 1

        row_map = {
            headers[idx]: (cell if cell is not None else "")
            for idx, cell in enumerate(row)
            if idx < len(headers)
        }

        short_name = str(row_map.get("商品简称") or "").strip()
        if not short_name:
            errors.append(f"第 {i} 行：商品简称为空，已跳过。")
            continue

        dto = {_COL_MAP[col]: str(row_map.get(col) or "").strip() for col in EXPECTED_HEADERS if col in headers}
        # 确保 short_name 已赋值（防止映射漏掉）
        dto["short_name"] = short_name
        if "short_title" in dto:
            dto["short_title"] = (dto.get("short_title") or "").strip()[
                :CART_SHORT_TITLE_MAX_LEN
            ]
        items.append(dto)
        success += 1

    failed = total - success
    return {
        "items": items,
        "total": total,
        "success": success,
        "failed": failed,
        "errors": errors,
    }
