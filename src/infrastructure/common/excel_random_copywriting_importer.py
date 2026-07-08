"""
文案库 Excel 导入工具
文件路径：src/infrastructure/common/excel_copywriting_importer.py
功能：解析标准模板 Excel，将文案数据转换为字典列表，供仓储层批量导入。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import List, Dict, Any
import logging

from openpyxl import load_workbook

from src.infrastructure.common.copywriting_work_id import (
    is_valid_copywriting_work_id,
"""
文案库 Excel 导入工具
文件路径：src/infrastructure/common/excel_copywriting_importer.py
功能：解析标准模板 Excel，将文案数据转换为字典列表，供仓储层批量导入。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import List, Dict, Any
import logging

from openpyxl import load_workbook

from src.infrastructure.common.copywriting_work_id import (
    is_valid_copywriting_work_id,
    COPYWRITING_WORK_ID_FORMAT_HINT,
)

logger = logging.getLogger(__name__)


# 宽松模式要求：只需能找到一列叫"文案"
DEFAULT_FIELD_ALIASES: Dict[str, List[str]] = {
    "work_id": ["作品编号", "编号", "work_id"],
    "short_title": ["作品标题", "标题", "title"],
    "description": ["作品描述", "描述", "备注", "description"],
    "content": ["作品文案", "文案内容", "内容", "文案", "content"],
}


def _detect_mapping(headers: List[str]) -> Dict[str, str]:
    mapping = {}
    for field_name, aliases in DEFAULT_FIELD_ALIASES.items():
        for alias in aliases:
            alias_lower = alias.strip().lower()
            for h in headers:
                if h and h.strip().lower() == alias_lower:
                    mapping[field_name] = h.strip()
                    break
            if field_name in mapping:
                break
    return mapping


@dataclass
class CopywritingDTO:
    work_id: str
    short_title: str
    description: str
    topics: str
    category: str
    content: str


def parse_excel(path: str) -> Dict[str, Any]:
    """解析随机文案库 Excel 文件。"""
    wb = load_workbook(filename=path, read_only=True, data_only=True)

    items: List[Dict[str, Any]] = []
    errors: List[str] = []
    total = 0
    success = 0
    total_sheets = 0
    valid_sheets = 0
    
    valid_sheet_found = False

    for sheet in wb.worksheets:
        total_sheets += 1
        sheet_has_items = False
        headers = []
        mapping = {}
        
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i == 1:
                headers = [str(c or "").strip() for c in row if str(c or "").strip()]
                mapping = _detect_mapping(headers)
                
                if "content" not in mapping:
                    err_msg = f"工作表「{sheet.title}」未找到「文案」列，已跳过"
                    errors.append(err_msg)
                    break
                
                valid_sheet_found = True
                continue

            if not any(row):
                continue

            # 构造列名 -> 值的映射
            row_map = {headers[idx]: (cell if cell is not None else "") for idx, cell in enumerate(row) if idx < len(headers)}

            # 获取各字段值，宽容模式
            content_col = mapping.get("content")
            content = str(row_map.get(content_col) or "").strip()
            
            if not content:
                continue

            work_id_col = mapping.get("work_id")
            work_id = str(row_map.get(work_id_col) or "").strip() if work_id_col else ""
            
            title_col = mapping.get("short_title")
            short_title = str(row_map.get(title_col) or "").strip() if title_col else ""
            
            desc_col = mapping.get("description")
            description = str(row_map.get(desc_col) or "").strip() if desc_col else ""

            total += 1

            items.append({
                "category": sheet.title,  # 用 Sheet 名作为分类
                "work_id": work_id,
                "short_title": short_title,
                "description": description,
                "content": content,
                "topics": "",
            })
            success += 1
            sheet_has_items = True
            
        if sheet_has_items:
            valid_sheets += 1

    wb.close()

    if not valid_sheet_found and not errors:
        errors.append("未在任何工作表中找到带有有效表头的数据。请确保存在一列名为「文案」或「内容」。")

    return {
        "items": items,
        "total": total,
        "success": success,
        "failed": total - success,
        "errors": errors,
        "total_sheets": total_sheets,
        "valid_sheets": valid_sheets
    }
