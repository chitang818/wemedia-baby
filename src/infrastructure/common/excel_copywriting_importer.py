"""
文案库 Excel 导入工具
文件路径：src/infrastructure/common/excel_copywriting_importer.py
功能：解析标准模板 Excel，将文案数据转换为字典列表，供仓储层批量导入。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import List, Dict, Any
import logging

from openpyxl import load_workbook

from src.infrastructure.common.copywriting_work_id import (
    is_valid_copywriting_work_id,
    COPYWRITING_WORK_ID_FORMAT_HINT,
)

logger = logging.getLogger(__name__)


# 严格模式要求的 4 列
EXACT_HEADERS = {"作品编号", "作品标题", "作品描述", "作品文案"}


@dataclass
class CopywritingDTO:
    work_id: str
    short_title: str
    description: str
    topics: str
    category: str
    content: str


def parse_excel(path: str, strict: bool = True) -> Dict[str, Any]:
    """解析文案库 Excel 文件。

    Args:
        path: Excel 文件路径。
        strict: 是否开启严格校验模式。标准文案库开启，随机文案库可关闭。

    Returns:
        dict: {
            "items": List[Dict[str, Any]],
            "total": int,
            "success": int,
            "failed": int,
            "errors": List[str],
        }
    """
    wb = load_workbook(filename=path, read_only=True, data_only=True)

    items: List[Dict[str, Any]] = []
    errors: List[str] = []
    total = 0
    success = 0
    total_sheets = 0
    valid_sheets = 0
    
    valid_sheet_found = False
    
    for sheet in wb.worksheets:
        if "使用说明" in sheet.title:
            continue
            
        total_sheets += 1
        sheet_has_items = False
        headers = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if i == 1:
                headers = [str(c or "").strip() for c in row if str(c or "").strip()]
                
                if not EXACT_HEADERS.issubset(set(headers)):
                    missing = EXACT_HEADERS - set(headers)
                    errors.append(f"工作表「{sheet.title}」被跳过：缺少必需列 {', '.join(missing)}")
                    break  # 跳过当前工作表
                
                valid_sheet_found = True
                continue

            if not any(row):
                continue

            # 构造列名 -> 值的映射
            row_map = {headers[idx]: (cell if cell is not None else "") for idx, cell in enumerate(row) if idx < len(headers)}

            # 获取各字段值，严格对应4列
            work_id = str(row_map.get("作品编号") or "").strip()
            short_title = str(row_map.get("作品标题") or "").strip()
            description = str(row_map.get("作品描述") or "").strip()
            content = str(row_map.get("作品文案") or "").strip()
            topics = ""

            total += 1

            # 校验逻辑
            if strict:
                # 严格模式：编号和描述必填，且编号格式必须正确
                if not work_id or not description:
                    errors.append(f"Sheet「{sheet.title}」第 {i} 行：作品编号或作品描述为空，已跳过。")
                    continue
                if not is_valid_copywriting_work_id(work_id):
                    errors.append(
                        f"Sheet「{sheet.title}」第 {i} 行：作品编号「{work_id}」不符合格式（{COPYWRITING_WORK_ID_FORMAT_HINT}），已跳过。"
                    )
                    continue
            else:
                # 非严格模式（随机文案库）：只要有作品描述即可
                if not description:
                    errors.append(f"Sheet「{sheet.title}」第 {i} 行：作品描述为空，已跳过。")
                    continue
                # 编号格式不规范仅记录警告，不阻断
                if work_id and not is_valid_copywriting_work_id(work_id):
                    errors.append(f"Sheet「{sheet.title}」第 {i} 行：作品编号「{work_id}」格式不规范，但已允许导入。")

            dto = {
                "work_id": work_id,
                "short_title": short_title,
                "description": description,
                "topics": topics,
                "category": sheet.title,
                "content": content,
            }
            items.append(dto)
            success += 1
            sheet_has_items = True

        if sheet_has_items:
            valid_sheets += 1

    if not valid_sheet_found:
        raise ValueError("Excel 中未找到任何包含有效表头的工作表。")

    failed = total - success
    return {
        "items": items,
        "total": total,
        "success": success,
        "failed": failed,
        "errors": errors,
        "total_sheets": total_sheets,
        "valid_sheets": valid_sheets,
    }
