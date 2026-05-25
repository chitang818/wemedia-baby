"""
位置推广 Excel 导入工具
功能：解析标准模板 Excel，将位置配置数据转换为字典列表，供仓储层批量导入。
"""

from __future__ import annotations

from typing import List, Dict, Any
import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

REQUIRED_HEADERS = ["位置简称"]
EXPECTED_HEADERS = [
    "位置简称",
    "抖音位置",
    "快手位置",
    "视频号位置",
    "小红书位置",
]

_COL_MAP = {
    "位置简称": "short_name",
    "抖音位置": "douyin_location",
    "快手位置": "kuaishou_location",
    "视频号位置": "channels_location",
    "小红书位置": "xiaohongshu_location",
}


def parse_excel(path: str) -> Dict[str, Any]:
    """解析位置推广 Excel 文件。"""
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = wb.active

    headers: List[str] = []
    items: List[Dict[str, Any]] = []
    errors: List[str] = []
    total = 0
    success = 0

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [str(c or "").strip() for c in row]
            missing = [h for h in REQUIRED_HEADERS if h not in headers]
            if missing:
                raise ValueError(f"Excel 模板表头缺失：{', '.join(missing)}")
            continue

        if not any(row):
            continue

        total += 1

        row_map = {
            headers[idx]: (cell if cell is not None else "")
            for idx, cell in enumerate(row)
            if idx < len(headers)
        }

        short_name = str(row_map.get("位置简称") or "").strip()
        if not short_name:
            errors.append(f"第 {i} 行：位置简称为空，已跳过。")
            continue

        dto = {
            _COL_MAP[col]: str(row_map.get(col) or "").strip()
            for col in EXPECTED_HEADERS
            if col in headers
        }
        dto["short_name"] = short_name
        items.append(dto)
        success += 1

    failed = total - success
    return {
        "items": items,
        "total": total,
        "success": success,
        "failed": failed,
        "errors": errors,
    }
